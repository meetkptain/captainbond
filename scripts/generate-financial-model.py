#!/usr/bin/env python3
"""Génère le modèle financier unique pour PLAN_STRATEGIQUE.md."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path("docs/superpowers/plans/PLAN_STRATEGIQUE_financial_model.xlsx")

# Hypothèses de base (scénario optimisé)
BASE_ASSUMPTIONS = {
    "Visiteurs landing M1": 2000,
    "Visiteurs landing M12": 100000,
    "Croissance visiteurs M1-M12": "linéaire",
    "Taux conversion visiteur → room": 0.12,
    "Taux completion 3 cartes gratuites": 0.75,
    "Taux conversion room → Pass 24h": 0.10,
    "Prix Pass 24h": 2.99,
    "Taux fin de partie → Dossier individuel": 0.20,
    "Prix Dossier individuel": 4.99,
    "Taux Dossier individuel → Dossier couple": 0.08,
    "Prix Dossier couple": 9.99,
    "Taux Dossier couple → Abonnement couple": 0.15,
    "Prix Abonnement mensuel": 7.99,
    "Churn mensuel abonnement": 0.05,
    "Clients B2B M6": 3,
    "Clients B2B M12": 15,
    "Clients B2B M24": 50,
    "ACV B2B moyen": 9600,
    "MRR B2B par client": 800,
    "Marketplace/White-label M12": 0,
    "Marketplace/White-label M24": 15000,
}


def style_header(ws, row):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_currency(ws, col_letters):
    for col in col_letters:
        for cell in ws[col][1:]:
            cell.number_format = '#,##0.00 €'


def style_percent(ws, col_letters):
    for col in col_letters:
        for cell in ws[col][1:]:
            cell.number_format = '0.00%'


def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def build_projection(
    assumptions,
    visitor_growth_factor=1.0,
    conversion_factor=1.0,
    b2b_delay_months=0,
    b2b_clients_factor=1.0,
    marketplace_m12=0,
    marketplace_m24=0,
):
    """Construit une projection mensuelle M1-M24 selon les hypothèses."""
    rows = []
    visitors_m1 = assumptions["Visiteurs landing M1"] * visitor_growth_factor
    visitors_m12 = assumptions["Visiteurs landing M12"] * visitor_growth_factor
    monthly_visitor_step = (visitors_m12 - visitors_m1) / 11

    mrr = 0.0
    total_subscribers = 0
    total_b2b_clients = 0

    for month in range(1, 25):
        if month <= 12:
            visitors = visitors_m1 + monthly_visitor_step * (month - 1)
        else:
            # Croissance plus lente M13-M24
            visitors = visitors_m12 * (1 + 0.05 * (month - 12))

        rooms = visitors * assumptions["Taux conversion visiteur → room"] * conversion_factor
        completions = rooms * assumptions["Taux completion 3 cartes gratuites"]
        pass_24h = completions * assumptions["Taux conversion room → Pass 24h"] * conversion_factor
        revenue_pass = pass_24h * assumptions["Prix Pass 24h"]

        dossier_ind = completions * assumptions["Taux fin de partie → Dossier individuel"] * conversion_factor
        revenue_dossier_ind = dossier_ind * assumptions["Prix Dossier individuel"]

        dossier_couple = dossier_ind * assumptions["Taux Dossier individuel → Dossier couple"] * conversion_factor
        revenue_dossier_couple = dossier_couple * assumptions["Prix Dossier couple"]

        new_subs = dossier_couple * assumptions["Taux Dossier couple → Abonnement couple"] * conversion_factor
        churned = total_subscribers * assumptions["Churn mensuel abonnement"]
        total_subscribers = max(0, total_subscribers + new_subs - churned)
        mrr = total_subscribers * assumptions["Prix Abonnement mensuel"]

        b2b_new = 0
        if month >= (6 + b2b_delay_months) and month < 12:
            b2b_new = (assumptions["Clients B2B M6"] / 6) * b2b_clients_factor
        elif month >= 12:
            b2b_new = (assumptions["Clients B2B M24"] / 12) * b2b_clients_factor
        total_b2b_clients += b2b_new
        revenue_b2b = total_b2b_clients * assumptions["MRR B2B par client"]

        marketplace = 0
        if month >= 12:
            marketplace = marketplace_m12 + (marketplace_m24 - marketplace_m12) * (month - 12) / 12

        revenue_b2c = revenue_pass + revenue_dossier_ind + revenue_dossier_couple + mrr
        revenue_total = revenue_b2c + revenue_b2b + marketplace

        rows.append({
            "Mois": month,
            "Visiteurs": visitors,
            "Rooms": rooms,
            "Completions": completions,
            "Pass 24h": pass_24h,
            "Revenus Pass 24h": revenue_pass,
            "Dossiers individuels": dossier_ind,
            "Revenus Dossiers ind.": revenue_dossier_ind,
            "Dossiers couple": dossier_couple,
            "Revenus Dossiers couple": revenue_dossier_couple,
            "Nouveaux abonnements": new_subs,
            "Abonnés actifs": total_subscribers,
            "MRR Couple": mrr,
            "Revenus B2C": revenue_b2c,
            "Nouveaux clients B2B": b2b_new,
            "Clients B2B cumulés": total_b2b_clients,
            "Revenus B2B": revenue_b2b,
            "Revenus Marketplace/White-label": marketplace,
            "Revenus total": revenue_total,
        })

    return rows


def write_assumptions_sheet(wb):
    ws = wb.active
    ws.title = "Hypothèses"
    ws.append(["Hypothèse", "Valeur", "Source / Note"])
    style_header(ws, 1)

    notes = {
        "Visiteurs landing M1": "SEO naissant, pas de paid ads",
        "Visiteurs landing M12": "SEO mature + affiliation + contenu",
        "Croissance visiteurs M1-M12": "Linéaire puis +5%/mois M13-M24",
        "Taux conversion visiteur → room": "Benchmark SaaS / jeu social",
        "Taux completion 3 cartes gratuites": "Mesuré après onboarding fix",
        "Taux conversion room → Pass 24h": "Objectif optimisé (jauge + paywall tour 2)",
        "Prix Pass 24h": "Catalogue actuel",
        "Taux fin de partie → Dossier individuel": "Upsell fin de soirée",
        "Prix Dossier individuel": "Catalogue actuel",
        "Taux Dossier individuel → Dossier couple": "Part des joueurs en couple ou curieux",
        "Prix Dossier couple": "Catalogue actuel",
        "Taux Dossier couple → Abonnement couple": "Trial 7 jours",
        "Prix Abonnement mensuel": "Catalogue actuel",
        "Churn mensuel abonnement": "Conservateur",
        "Clients B2B M6": "Prospection manuelle bars/cafés",
        "Clients B2B M12": "Case studies + LinkedIn",
        "Clients B2B M24": "Outbound automatisé",
        "ACV B2B moyen": "Entre Pack Événement 399€ et Premium 799€",
        "Marketplace/White-label M12": "Non prioritaire M12",
        "Marketplace/White-label M24": "White label + marketplace créateurs",
    }

    for key, value in BASE_ASSUMPTIONS.items():
        ws.append([key, value, notes.get(key, "")])

    autosize_columns(ws)


def write_projection_sheet(wb, title, rows, assumptions):
    ws = wb.create_sheet(title)
    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws, 1)

    for row in rows:
        ws.append([row[h] for h in headers])

    # Formats
    currency_cols = []
    percent_cols = []
    for idx, h in enumerate(headers, 1):
        if "Taux" in h or "%" in h or "Churn" in h:
            percent_cols.append(get_column_letter(idx))
        elif "Revenu" in h or "Prix" in h or "ACV" in h or "MRR" in h:
            currency_cols.append(get_column_letter(idx))

    style_currency(ws, currency_cols)
    style_percent(ws, percent_cols)

    # Ligne de synthèse M3/M6/M12/M24
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value="Synthèse")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    milestones = [3, 6, 12, 18, 24]
    for i, m in enumerate(milestones):
        row_idx = m - 1
        r = rows[row_idx]
        ws.cell(row=summary_row + 1 + i, column=1, value=f"M{m}")
        ws.cell(row=summary_row + 1 + i, column=2, value=r["Revenus B2C"])
        ws.cell(row=summary_row + 1 + i, column=3, value=r["Revenus B2B"])
        ws.cell(row=summary_row + 1 + i, column=4, value=r["Revenus Marketplace/White-label"])
        ws.cell(row=summary_row + 1 + i, column=5, value=r["Revenus total"])
        ws.cell(row=summary_row + 1 + i, column=6, value=r["Abonnés actifs"])

    for col in ["B", "C", "D", "E"]:
        for cell in ws[col][summary_row:summary_row + len(milestones)]:
            cell.number_format = '#,##0.00 €'

    ws.cell(row=summary_row, column=2, value="B2C")
    ws.cell(row=summary_row, column=3, value="B2B")
    ws.cell(row=summary_row, column=4, value="Marketplace")
    ws.cell(row=summary_row, column=5, value="Total")
    ws.cell(row=summary_row, column=6, value="Abonnés")
    for col in range(1, 7):
        ws.cell(row=summary_row, column=col).font = Font(bold=True)

    autosize_columns(ws)


def main():
    wb = openpyxl.Workbook()
    write_assumptions_sheet(wb)

    # Scénario Optimisé
    optimized = build_projection(BASE_ASSUMPTIONS)
    write_projection_sheet(wb, "Optimisé", optimized, BASE_ASSUMPTIONS)

    # Scénario Conservateur
    conservative = build_projection(
        BASE_ASSUMPTIONS,
        visitor_growth_factor=0.6,
        conversion_factor=0.6,
        b2b_delay_months=6,
        b2b_clients_factor=0.5,
        marketplace_m12=0,
        marketplace_m24=2000,
    )
    write_projection_sheet(wb, "Conservateur", conservative, BASE_ASSUMPTIONS)

    # Scénario Moonshot
    moonshot = build_projection(
        BASE_ASSUMPTIONS,
        visitor_growth_factor=1.8,
        conversion_factor=1.4,
        b2b_delay_months=-3,
        marketplace_m12=3000,
        marketplace_m24=15000,
    )
    write_projection_sheet(wb, "Moonshot", moonshot, BASE_ASSUMPTIONS)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Modèle financier généré : {OUTPUT.resolve()}")


if __name__ == "__main__":
    main()
